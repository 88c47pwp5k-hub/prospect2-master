"""
GÉNÉRATEUR DE PRODUCTION — CADRE MURRET
Solarium Pro | Avec gestion des restes de barres

UTILISATION:
    python generateur_cadre_avec_restes.py

DÉPENDANCES:
    pip install reportlab openpyxl

FICHIER EXCEL:
    Le script lit et écrit dans INVENTAIRE_MASTER.xlsx, onglet RESTES.
    Si le fichier n'existe pas, les restes sont ignorés.
"""

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from fractions import Fraction
import math
import os
import json
from datetime import date

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
EXCEL_FILE = "INVENTAIRE_MASTER.xlsx"   # Chemin vers ton fichier Excel
ONGLET_RESTES = "RESTES"
BARRE_LONGUEUR = 240       # pouces — barre standard 240"
KERF = 0.25                # largeur de coupe en pouces
BOUT_CHAQUE = 1.0          # réservé à chaque extrémité de barre
LONGUEUR_UTILE = BARRE_LONGUEUR - (2 * BOUT_CHAQUE)  # 238"
SEUIL_RESTE_MIN = 6        # ne pas sauvegarder les restes < 6 pouces

# Couleurs Solarium Pro
BLEU = colors.HexColor("#1F3864")
OR   = colors.HexColor("#C68B00")
BLEU_CLAIR = colors.HexColor("#D6E4F0")
OR_CLAIR   = colors.HexColor("#FFF3CD")
GRIS_CLAIR = colors.HexColor("#F5F5F5")
VERT_CLAIR = colors.HexColor("#D4EDDA")


# ─── CONVERSION POUCES ────────────────────────────────────────────────────────
def fraction_vers_decimal(valeur_str):
    """Convertit '39-1/2' ou '39 1/2' → 39.5, '194-5/8' ou '194 5/8' → 194.625"""
    valeur_str = valeur_str.strip().replace('"', '')
    if '-' in valeur_str:
        parties = valeur_str.split('-', 1)
        entier = float(parties[0])
        fraction = float(Fraction(parties[1]))
        return entier + fraction
    elif ' ' in valeur_str and '/' in valeur_str:
        parties = valeur_str.split(' ', 1)
        entier = float(parties[0])
        fraction = float(Fraction(parties[1]))
        return entier + fraction
    elif '/' in valeur_str:
        return float(Fraction(valeur_str))
    else:
        return float(valeur_str)

def decimal_vers_fraction(decimal):
    """Convertit 39.5 → '39-1/2" pour affichage — arrondi au 1/16" le plus proche.
    Fraction(reste).limit_denominator(16) produisait des dénominateurs non-standard
    (4/7, 5/14…) car Python choisit la meilleure approximation rationnelle ≤16,
    pas nécessairement une puissance de 2. On arrondit d'abord au 1/16 exact."""
    from math import gcd
    seizieme = round(decimal * 16) / 16
    entier = int(seizieme)
    reste = seizieme - entier
    if reste < 0.001:
        return f'{entier}"'
    num = round(reste * 16)
    den = 16
    g = gcd(num, den)
    num //= g; den //= g
    if entier == 0:
        return f'{num}/{den}"'
    return f'{entier}-{num}/{den}"'


# ─── CALCUL DES PIÈCES ────────────────────────────────────────────────────────
def calculer_pieces(largeur_str, hauteur_str, nb_montants):
    """
    Retourne la liste des pièces à couper pour chaque type de profil.
    Chaque pièce = dict {nom, longueur_dec, longueur_str, qte}
    """
    largeur = fraction_vers_decimal(largeur_str)
    hauteur = fraction_vers_decimal(hauteur_str)
    nb_sections = nb_montants - 1

    # Longueurs déduites
    traverse_haut = largeur
    montant_vertical = hauteur - 2.5          # déduction standard 2-1/2"
    traverse_bas = (largeur - (nb_montants * 2.5)) / nb_sections  # entre montants

    pieces = [
        {"nom": "Traverse du haut",   "longueur_dec": traverse_haut,     "longueur_str": decimal_vers_fraction(traverse_haut),    "qte": 1},
        {"nom": "Montants verticaux", "longueur_dec": montant_vertical,   "longueur_str": decimal_vers_fraction(montant_vertical),  "qte": nb_montants},
        {"nom": "Traverses du bas",   "longueur_dec": traverse_bas,       "longueur_str": decimal_vers_fraction(traverse_bas),      "qte": nb_sections},
    ]
    return pieces


# ─── LECTURE DES RESTES EXCEL ─────────────────────────────────────────────────
def lire_restes_disponibles(couleur_projet, type_profil_filtre=None):
    """
    Lit l'onglet RESTES de l'Excel et retourne les restes DISPONIBLES
    correspondant à la couleur du projet.
    Retourne: liste de dicts triés par longueur décroissante
    """
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            return []
        wb = load_workbook(EXCEL_FILE, data_only=True)
        if ONGLET_RESTES not in wb.sheetnames:
            return []
        ws = wb[ONGLET_RESTES]

        restes = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            if not row[0]:  # ID vide = fin du tableau
                continue
            id_reste, no_projet, client, type_profil, longueur, couleur, date_ajout, statut, no_utilise, notes = row[:10]
            if statut == "DISPONIBLE" and couleur == couleur_projet:
                if type_profil_filtre is None or type_profil == type_profil_filtre:
                    restes.append({
                        "id": id_reste,
                        "no_projet": no_projet,
                        "type_profil": type_profil,
                        "longueur": float(longueur) if longueur else 0,
                        "couleur": couleur,
                        "utilise": False,
                    })
        # Trier du plus long au plus court (First Fit Decreasing)
        return sorted(restes, key=lambda x: x["longueur"], reverse=True)
    except Exception as e:
        print(f"  Avertissement: impossible de lire les restes Excel ({e})")
        return []


def marquer_restes_utilises(restes_utilises, no_projet_nouveau):
    """Met à jour l'Excel pour marquer les restes consommés comme UTILISÉS"""
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            return
        wb = load_workbook(EXCEL_FILE)
        if ONGLET_RESTES not in wb.sheetnames:
            return
        ws = wb[ONGLET_RESTES]
        ids_a_marquer = {r["id"] for r in restes_utilises}
        for row in ws.iter_rows(min_row=8):
            if row[0].value in ids_a_marquer:
                row[7].value = "UTILISÉ"       # colonne H = statut
                row[8].value = no_projet_nouveau  # colonne I = no projet utilisé
        wb.save(EXCEL_FILE)
        print(f"  ✓ {len(ids_a_marquer)} reste(s) marqué(s) UTILISÉ dans Excel")
    except Exception as e:
        print(f"  Avertissement: impossible de mettre à jour les restes ({e})")


def ajouter_nouveaux_restes(nouveaux_restes, no_projet, client, couleur):
    """Ajoute les nouveaux restes générés par ce projet dans l'Excel"""
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            print("  Info: INVENTAIRE_MASTER.xlsx introuvable — restes non sauvegardés")
            return
        wb = load_workbook(EXCEL_FILE)
        if ONGLET_RESTES not in wb.sheetnames:
            print(f"  Info: onglet '{ONGLET_RESTES}' introuvable — restes non sauvegardés")
            return
        ws = wb[ONGLET_RESTES]
        # Trouver la prochaine ligne vide (après row 7 = headers)
        prochaine_ligne = 8
        for row in ws.iter_rows(min_row=8, max_col=1):
            if row[0].value:
                prochaine_ligne += 1
        # Compter les IDs existants pour générer le prochain
        nb_existants = prochaine_ligne - 8
        for i, reste in enumerate(nouveaux_restes):
            id_reste = f"R-{nb_existants + i + 1:03d}"
            ws.cell(prochaine_ligne + i, 1).value = id_reste
            ws.cell(prochaine_ligne + i, 2).value = no_projet
            ws.cell(prochaine_ligne + i, 3).value = client
            ws.cell(prochaine_ligne + i, 4).value = reste["type_profil"]
            ws.cell(prochaine_ligne + i, 5).value = round(reste["longueur"], 4)
            ws.cell(prochaine_ligne + i, 6).value = couleur
            ws.cell(prochaine_ligne + i, 7).value = str(date.today())
            ws.cell(prochaine_ligne + i, 8).value = "DISPONIBLE"
            ws.cell(prochaine_ligne + i, 9).value = ""
            ws.cell(prochaine_ligne + i, 10).value = reste.get("notes", "")
        wb.save(EXCEL_FILE)
        print(f"  ✓ {len(nouveaux_restes)} nouveau(x) reste(s) ajouté(s) dans Excel")
    except Exception as e:
        print(f"  Avertissement: impossible d'ajouter les nouveaux restes ({e})")


# ─── ALGORITHME D'OPTIMISATION AVEC RESTES ────────────────────────────────────
def optimiser_coupes(pieces, type_profil, couleur, verbose=True):
    """
    First Fit Decreasing avec utilisation prioritaire des restes existants.
    Retourne: (barres_utilisées, restes_consommés, nouveaux_restes)
    """
    # 1. Charger les restes disponibles compatibles
    restes_dispo = lire_restes_disponibles(couleur, type_profil)
    if verbose and restes_dispo:
        print(f"\n  → {len(restes_dispo)} reste(s) disponible(s) pour {type_profil} ({couleur})")
        for r in restes_dispo:
            print(f"     {r['id']}: {decimal_vers_fraction(r['longueur'])} (projet {r['no_projet']})")

    # 2. Aplatir la liste des pièces (une entrée par pièce individuelle)
    toutes_pieces = []
    for p in pieces:
        for _ in range(p["qte"]):
            toutes_pieces.append({"nom": p["nom"], "longueur": p["longueur_dec"]})
    # Trier du plus long au plus court
    toutes_pieces.sort(key=lambda x: x["longueur"], reverse=True)

    # 3. Construire la liste des "contenants" disponibles
    #    D'abord les restes, ensuite des barres neuves si nécessaire
    contenants = []
    for r in restes_dispo:
        contenants.append({
            "type": "reste",
            "id_reste": r["id"],
            "longueur_totale": r["longueur"],
            "pieces": [],
            "utilise_dec": 0,
            "reste_final": r["longueur"],
            "reste_obj": r,
        })

    restes_consommes = []
    nouveaux_restes = []
    barres_neuves = []

    # 4. First Fit Decreasing — placer chaque pièce dans le premier contenant qui a la place
    for piece in toutes_pieces:
        longueur_necessaire = piece["longueur"] + KERF
        placee = False

        for c in contenants:
            espace_restant = c["longueur_totale"] - c["utilise_dec"]
            if c["type"] == "barre":
                # Pour les barres neuves, tenir compte du kerf initial
                espace_restant = LONGUEUR_UTILE - c["utilise_dec"]
            if espace_restant >= longueur_necessaire:
                c["pieces"].append(piece)
                c["utilise_dec"] += longueur_necessaire
                if c["type"] == "reste":
                    c["reste_final"] = c["longueur_totale"] - c["utilise_dec"]
                    if not c.get("marque_utilise") and c["utilise_dec"] > 0:
                        c["marque_utilise"] = True
                placee = True
                break

        if not placee:
            # Ouvrir une nouvelle barre
            nouvelle_barre = {
                "type": "barre",
                "num": len([c for c in contenants if c["type"] == "barre"]) + 1,
                "longueur_totale": BARRE_LONGUEUR,
                "pieces": [piece],
                "utilise_dec": longueur_necessaire,
                "reste_final": LONGUEUR_UTILE - longueur_necessaire,
            }
            contenants.append(nouvelle_barre)

    # 5. Séparer résultats
    for c in contenants:
        if c["type"] == "reste" and c["utilise_dec"] > 0:
            restes_consommes.append(c["reste_obj"])
            if c["reste_final"] >= SEUIL_RESTE_MIN:
                nouveaux_restes.append({
                    "type_profil": type_profil,
                    "longueur": c["reste_final"],
                    "notes": f"Reste après utilisation de {c['id_reste']}",
                })
        elif c["type"] == "barre":
            barres_neuves.append(c)
            if c["reste_final"] >= SEUIL_RESTE_MIN:
                nouveaux_restes.append({
                    "type_profil": type_profil,
                    "longueur": c["reste_final"],
                    "notes": f"Barre #{c['num']} — reste après production",
                })

    return contenants, restes_consommes, nouveaux_restes


# ─── GÉNÉRATION PDF ───────────────────────────────────────────────────────────
def dessiner_entete(c_pdf, titre_doc, no_projet, client, largeur, hauteur_page):
    """Dessine l'en-tête Solarium Pro"""
    # Bande bleue
    c_pdf.setFillColor(BLEU)
    c_pdf.rect(0, hauteur_page - 1.1*inch, 8.5*inch, 1.1*inch, fill=True, stroke=False)

    # Titre document
    c_pdf.setFillColor(OR)
    c_pdf.setFont("Helvetica-Bold", 16)
    c_pdf.drawString(0.4*inch, hauteur_page - 0.45*inch, titre_doc)

    # Sous-titre
    c_pdf.setFillColor(colors.white)
    c_pdf.setFont("Helvetica", 9)
    c_pdf.drawString(0.4*inch, hauteur_page - 0.72*inch, f"No: {no_projet}  |  Client: {client}")

    # Date
    c_pdf.drawRightString(8.1*inch, hauteur_page - 0.72*inch, str(date.today()))


def dessiner_ligne_info(c_pdf, texte, y, fond=None):
    """Ligne de résumé colorée"""
    if fond:
        c_pdf.setFillColor(fond)
        c_pdf.rect(0.4*inch, y - 2, 7.7*inch, 16, fill=True, stroke=False)
    c_pdf.setFillColor(BLEU)
    c_pdf.setFont("Helvetica", 9)
    c_pdf.drawString(0.5*inch, y + 2, texte)


def generer_pdf_production(params, contenants_par_profil, restes_consommes, nouveaux_restes, fichier_sortie):
    """Génère le PDF de production complet avec section restes"""
    no_projet = params["no_projet"]
    client    = params["client"]
    couleur   = params["couleur"]

    c_pdf = canvas.Canvas(fichier_sortie, pagesize=letter)
    w, h = letter
    y = h - 1.3*inch

    dessiner_entete(c_pdf, "DOCUMENT DE PRODUCTION — Cadre murret", no_projet, client, w, h)

    # ── Bandeau récap ──
    c_pdf.setFillColor(BLEU_CLAIR)
    c_pdf.rect(0.4*inch, y - 0.3*inch, 7.7*inch, 0.3*inch, fill=True, stroke=False)
    c_pdf.setFillColor(BLEU)
    c_pdf.setFont("Helvetica-Bold", 9)
    c_pdf.drawString(0.5*inch, y - 0.18*inch,
        f"{params['largeur']} x {params['hauteur']}  |  {params['nb_montants']} montants  |  "
        f"{params['nb_montants']-1} sections  |  Tube alu 2-1/2\"  |  Couleur: {couleur}")
    y -= 0.55*inch

    # ── SECTION RESTES UTILISÉS ──
    if restes_consommes:
        c_pdf.setFillColor(VERT_CLAIR)
        c_pdf.rect(0.4*inch, y - 0.25*inch, 7.7*inch, 0.25*inch, fill=True, stroke=False)
        c_pdf.setFillColor(colors.HexColor("#155724"))
        c_pdf.setFont("Helvetica-Bold", 10)
        c_pdf.drawString(0.5*inch, y - 0.12*inch, f"✓ RESTES RÉUTILISÉS ({len(restes_consommes)} barre(s) récupérée(s))")
        y -= 0.35*inch
        c_pdf.setFont("Helvetica", 8)
        c_pdf.setFillColor(BLEU)
        for r in restes_consommes:
            c_pdf.drawString(0.6*inch, y, f"  • {r['id']} — {r['type_profil']} — {decimal_vers_fraction(r['longueur'])} ({r['couleur']}) — projet {r['no_projet']}")
            y -= 0.2*inch
        y -= 0.1*inch

    # ── TABLEAUX DE COUPE PAR PROFIL ──
    for type_profil, contenants in contenants_par_profil.items():
        # Titre section
        c_pdf.setFillColor(BLEU)
        c_pdf.rect(0.4*inch, y - 0.22*inch, 7.7*inch, 0.22*inch, fill=True, stroke=False)
        c_pdf.setFillColor(colors.white)
        c_pdf.setFont("Helvetica-Bold", 10)
        c_pdf.drawString(0.5*inch, y - 0.14*inch, f"{type_profil.upper()} — Barres de 240\"")
        y -= 0.35*inch

        barres_neuves = [c for c in contenants if c["type"] == "barre"]
        restes_utilises_ici = [c for c in contenants if c["type"] == "reste" and c["utilise_dec"] > 0]

        c_pdf.setFont("Helvetica", 9)
        c_pdf.setFillColor(BLEU)

        # Résumé barres
        nb_neuves = len(barres_neuves)
        nb_recup  = len(restes_utilises_ici)
        c_pdf.drawString(0.5*inch, y, f"Barres neuves requises: {nb_neuves}   |   Restes réutilisés: {nb_recup}")
        y -= 0.25*inch

        # Tableau des barres neuves
        if barres_neuves:
            data = [["#", "Source", "Découpes", "Utilisé", "Reste"]]
            for c_barre in barres_neuves:
                decoupe_str = " + ".join([decimal_vers_fraction(p["longueur"]) for p in c_barre["pieces"]])
                utilise_str = decimal_vers_fraction(c_barre["utilise_dec"])
                reste_str   = decimal_vers_fraction(c_barre["reste_final"]) if c_barre["reste_final"] > 0 else "—"
                data.append([f"#{c_barre['num']}", "Barre neuve 240\"", decoupe_str, utilise_str, reste_str])

            t = Table(data, colWidths=[0.4*inch, 1.2*inch, 4.0*inch, 0.9*inch, 0.9*inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), BLEU),
                ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("BACKGROUND", (0,1), (-1,-1), GRIS_CLAIR),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, GRIS_CLAIR]),
                ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
                ("ALIGN",      (3,0), (-1,-1), "CENTER"),
                ("ALIGN",      (0,0), (1,-1), "CENTER"),
            ]))
            t.wrapOn(c_pdf, w, h)
            t_w, t_h = t.wrap(0, 0)
            if y - t_h < 0.8*inch:
                c_pdf.showPage()
                y = h - 0.5*inch
            t.drawOn(c_pdf, 0.4*inch, y - t_h)
            y -= t_h + 0.15*inch

        # Restes utilisés pour ce profil
        if restes_utilises_ici:
            data2 = [["#", "ID Reste", "Source", "Découpes", "Reste final"]]
            for c_reste in restes_utilises_ici:
                decoupe_str = " + ".join([decimal_vers_fraction(p["longueur"]) for p in c_reste["pieces"]])
                reste_str   = decimal_vers_fraction(c_reste["reste_final"]) if c_reste["reste_final"] > 0 else "Épuisé"
                data2.append(["R", c_reste["id_reste"], f"Projet {c_reste['reste_obj']['no_projet']}", decoupe_str, reste_str])

            t2 = Table(data2, colWidths=[0.4*inch, 0.9*inch, 1.3*inch, 3.5*inch, 1.0*inch])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#155724")),
                ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("BACKGROUND", (0,1), (-1,-1), VERT_CLAIR),
                ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
                ("ALIGN",      (0,0), (1,-1), "CENTER"),
                ("ALIGN",      (4,0), (4,-1), "CENTER"),
            ]))
            t2.wrapOn(c_pdf, w, h)
            t2_w, t2_h = t2.wrap(0, 0)
            if y - t2_h < 0.8*inch:
                c_pdf.showPage()
                y = h - 0.5*inch
            t2.drawOn(c_pdf, 0.4*inch, y - t2_h)
            y -= t2_h + 0.25*inch

    # ── NOUVEAUX RESTES GÉNÉRÉS ──
    if nouveaux_restes:
        if y < 2.5*inch:
            c_pdf.showPage()
            y = h - 0.5*inch

        c_pdf.setFillColor(OR_CLAIR)
        c_pdf.rect(0.4*inch, y - 0.22*inch, 7.7*inch, 0.22*inch, fill=True, stroke=False)
        c_pdf.setFillColor(OR)
        c_pdf.setFont("Helvetica-Bold", 10)
        c_pdf.drawString(0.5*inch, y - 0.14*inch, f"NOUVEAUX RESTES À ENREGISTRER ({len(nouveaux_restes)} pcs) — Ajoutés automatiquement dans INVENTAIRE_MASTER")
        y -= 0.35*inch

        data3 = [["Type Profil", "Longueur", "Couleur", "Notes"]]
        for nr in nouveaux_restes:
            data3.append([nr["type_profil"], decimal_vers_fraction(nr["longueur"]), couleur, nr.get("notes", "")])

        t3 = Table(data3, colWidths=[2.0*inch, 1.2*inch, 1.5*inch, 3.0*inch])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), OR),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("BACKGROUND", (0,1), (-1,-1), OR_CLAIR),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ]))
        t3.wrapOn(c_pdf, w, h)
        t3_w, t3_h = t3.wrap(0, 0)
        t3.drawOn(c_pdf, 0.4*inch, y - t3_h)
        y -= t3_h + 0.3*inch

    c_pdf.save()
    print(f"\n✓ PDF généré: {fichier_sortie}")


# ─── PROGRAMME PRINCIPAL ──────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  GÉNÉRATEUR DE PRODUCTION — CADRE MURRET")
    print("  Solarium Pro | Avec gestion des restes")
    print("=" * 60)

    # ── SAISIE DES PARAMÈTRES ──
    print("\nDimensions du cadre murret:")
    largeur_str    = input("  Largeur totale (ex: 194-5/8): ").strip() or "194-5/8"
    hauteur_str    = input("  Hauteur (ex: 42): ").strip() or "42"
    nb_montants    = int(input("  Nombre de montants verticaux (ex: 5): ").strip() or "5")
    no_projet      = input("  Numéro de projet (ex: 2024-002): ").strip() or "2024-002"
    client         = input("  Nom du client: ").strip() or "Client Test"
    couleur        = input("  Couleur (ex: Blanc cassé, Noir mat): ").strip() or "Blanc cassé"

    print(f"\nProjet: {no_projet} | Client: {client} | Couleur: {couleur}")
    print(f"Dimensions: {largeur_str}\" x {hauteur_str}\" | {nb_montants} montants\n")

    # ── CALCUL DES PIÈCES ──
    pieces = calculer_pieces(largeur_str, hauteur_str, nb_montants)

    # ── OPTIMISATION AVEC RESTES pour chaque profil ──
    profils = ["Tube alu 2-1/2\"", "Plaque pression", "Couvercle"]
    contenants_par_profil = {}
    tous_restes_consommes = []
    tous_nouveaux_restes  = []

    for profil in profils:
        print(f"Optimisation: {profil}...")
        contenants, restes_consommes, nouveaux_restes = optimiser_coupes(pieces, profil, couleur)
        contenants_par_profil[profil] = contenants
        tous_restes_consommes.extend(restes_consommes)
        tous_nouveaux_restes.extend(nouveaux_restes)

    # ── RÉSUMÉ CONSOLE ──
    print("\n── RÉSUMÉ ──────────────────────────────────────────")
    nb_barres_neuves = sum(len([c for c in ct if c["type"] == "barre"]) for ct in contenants_par_profil.values())
    nb_restes_utilises = len(tous_restes_consommes)
    nb_nouveaux = len(tous_nouveaux_restes)
    print(f"  Barres neuves ouvertes : {nb_barres_neuves}")
    print(f"  Restes existants utilisés : {nb_restes_utilises}")
    print(f"  Nouveaux restes créés : {nb_nouveaux}")
    if nb_restes_utilises > 0:
        print(f"  ★ ÉCONOMIE : {nb_restes_utilises} barre(s) récupérée(s) = pas de gaspillage!")

    # ── MISE À JOUR EXCEL ──
    print("\nMise à jour Excel...")
    if tous_restes_consommes:
        marquer_restes_utilises(tous_restes_consommes, no_projet)
    if tous_nouveaux_restes:
        ajouter_nouveaux_restes(tous_nouveaux_restes, no_projet, client, couleur)

    # ── GÉNÉRATION PDF ──
    params = {
        "no_projet": no_projet, "client": client, "couleur": couleur,
        "largeur": largeur_str + '"', "hauteur": hauteur_str + '"',
        "nb_montants": nb_montants,
    }
    fichier_sortie = f"production_{no_projet.replace('-','_')}_{client.replace(' ','_')}.pdf"
    generer_pdf_production(params, contenants_par_profil, tous_restes_consommes, tous_nouveaux_restes, fichier_sortie)

    print(f"\n✓ Terminé! Fichier: {fichier_sortie}")
    print("=" * 60)


if __name__ == "__main__":
    main()
