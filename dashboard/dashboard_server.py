#!/usr/bin/env python3
"""
Serveur Flask — Dashboard Solarium Pro
Port 7373 | Deux onglets : Leads + Soumissions
"""

import json, os, subprocess, sys
from datetime import datetime
from flask import Flask, jsonify, request, send_file

CONFIG = json.load(open(os.path.expanduser("~/prospect2-master/config.json")))

DOSSIER          = os.path.join(CONFIG["chemins"]["base"], CONFIG["chemins"]["dashboard"])
DASHBOARD_HTML   = os.path.join(DOSSIER, "dashboard_leads.html")
LEADS_JSON       = os.path.join(DOSSIER, "leads_dashboard.json")
LEADS_XLSX       = os.path.join(DOSSIER, "Leads_2026.xlsx")
SOUMISSIONS_JSON = os.path.join(DOSSIER, "soumissions.json")
SOUMISSIONS_JS   = os.path.join(DOSSIER, "soumissions.js")
SCANNER_PY       = os.path.join(DOSSIER, "scanner_soumissions.py")
GENERER_PY       = os.path.join(DOSSIER, "generer_dashboard.py")

app = Flask(__name__)

@app.after_request
def cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return response

@app.route("/", defaults={"path": ""}, methods=["OPTIONS"])
@app.route("/<path:path>", methods=["OPTIONS"])
def options(path):
    return "", 204


# ── Utilitaires ──────────────────────────────────────────────────────────────

def load_json(path):
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def save_soumissions(data):
    save_json(SOUMISSIONS_JSON, data)
    js = "var _SOUM_DATA = " + json.dumps(data, ensure_ascii=False) + ";"
    with open(SOUMISSIONS_JS, "w", encoding="utf-8") as f:
        f.write(js)

def regenerer_dashboard():
    subprocess.Popen(
        [sys.executable, GENERER_PY],
        stdout=open("/tmp/gen_dashboard.log", "w"),
        stderr=subprocess.STDOUT,
    )

def update_xlsx_statut(lead_id, statut):
    """Met à jour le statut d'un lead dans Leads_2026.xlsx."""
    if not os.path.exists(LEADS_XLSX):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(LEADS_XLSX)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=False))
        header_idx = next(
            (i for i, r in enumerate(rows) if r and r[0].value == "Nom"), None
        )
        if header_idx is None:
            return
        # Colonne 8 = statut (index 7, 1-based col 8)
        for row in ws.iter_rows(min_row=header_idx + 2):
            # On identifie la ligne par le nom + date stockés dans le JSON
            pass
        # Approche simple : réécrire toute la colonne statut depuis le JSON
        leads = load_json(LEADS_JSON)
        id_to_statut = {l["id"]: l["statut"] for l in leads}
        # La correspondance ID → ligne n'est pas stockée dans l'xlsx ;
        # on utilise leads_dashboard.json comme source de vérité et on
        # réécrit l'xlsx complet.
        from import_leads_v4 import lire_xlsx, sauvegarder_xlsx  # type: ignore
        _, leads_xlsx = lire_xlsx(LEADS_XLSX)
        # Mise à jour du statut dans la liste xlsx
        for l in leads_xlsx:
            match = next((j for j in leads if j["nom"] == l["nom"]), None)
            if match:
                l["statut"] = match["statut"]
        sauvegarder_xlsx(LEADS_XLSX, [], leads_xlsx)
    except Exception as e:
        print(f"Avertissement xlsx: {e}", file=sys.stderr)


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_file(DASHBOARD_HTML)


@app.route("/soumissions", methods=["GET"])
def get_soumissions():
    data = load_json(SOUMISSIONS_JSON)
    return jsonify(data)


@app.route("/scanner", methods=["POST"])
def scanner():
    try:
        result = subprocess.run(
            [sys.executable, SCANNER_PY],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            return jsonify({"ok": False, "error": result.stderr.strip()}), 500
        output = result.stdout.strip()
        # Le scanner écrit du JSON en dernière ligne
        last_line = output.splitlines()[-1] if output else "{}"
        data = json.loads(last_line)
        return jsonify(data)
    except subprocess.TimeoutExpired:
        return jsonify({"ok": False, "error": "Timeout IMAP"}), 504
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/update-statut", methods=["POST"])
def update_statut():
    body = request.get_json(silent=True) or {}
    lead_id = body.get("id")
    statut  = body.get("statut", "").strip()
    if not lead_id or not statut:
        return jsonify({"ok": False, "error": "Paramètres manquants"}), 400

    leads = load_json(LEADS_JSON)
    updated = False
    for l in leads:
        if l.get("id") == lead_id:
            l["statut"] = statut
            updated = True
            break

    if not updated:
        return jsonify({"ok": False, "error": "Lead introuvable"}), 404

    save_json(LEADS_JSON, leads)
    update_xlsx_statut(lead_id, statut)
    regenerer_dashboard()
    return jsonify({"ok": True})


@app.route("/update-soumission-statut", methods=["POST"])
def update_soumission_statut():
    body   = request.get_json(silent=True) or {}
    msg_id = body.get("msg_id", "").strip()
    statut = body.get("statut", "").strip()
    if not msg_id or not statut:
        return jsonify({"ok": False, "error": "Paramètres manquants"}), 400

    data = load_json(SOUMISSIONS_JSON)
    updated = False
    for s in data:
        if s.get("msg_id") == msg_id:
            s["statut"]      = statut
            s["date_statut"] = datetime.now().isoformat()
            updated = True
            break

    if not updated:
        return jsonify({"ok": False, "error": "Soumission introuvable"}), 404

    save_soumissions(data)
    return jsonify({"ok": True})


# ── Démarrage ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    PORT = CONFIG["serveurs"]["dashboard_port"]
    print(f"Dashboard Solarium Pro → http://localhost:{PORT}")
    app.run(host="127.0.0.1", port=PORT, debug=False)
